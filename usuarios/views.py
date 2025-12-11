from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import check_password, make_password
from .models import Usuario
from .forms import LoginForm, RegistroUsuarioForm, EditarPerfilForm, CargaMasivausuariosForm
from django.shortcuts import render
from .decorators import login_required, admin_required, funcionario_required
import pandas as pd
from centros.models import Centro
from soporte.models import Soporte
from coordinaciones.models import Coordinacion
from datetime import datetime, timedelta
from io import BytesIO
from django.http import HttpResponse
from django.db.models import Q, Count
import csv
from io import StringIO, BytesIO
from django.contrib.auth.hashers import make_password


try:
    from ambientes.models import Ambiente
except ImportError:
    Ambiente = None

try:
    from reservas.models import Reserva
except ImportError:
    Reserva = None


def login_view(request):
    if 'user_id' in request.session:
        if request.session['user_type'] == 'admin':
            return redirect('panel_admin')
        else:
            return redirect('panel_funcionario')
            
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            numero_documento = form.cleaned_data['numero_documento']
            password = form.cleaned_data['password']
            
            try:
                usuario = Usuario.objects.get(id=numero_documento)
                
                if check_password(password, usuario.password):
                    request.session['user_id'] = usuario.id
                    request.session['user_name'] = usuario.nombre
                    request.session['user_type'] = usuario.tipo
                    
                    messages.success(request, f'¡Bienvenido {usuario.nombre}!')
                    
                    if usuario.tipo == 'admin':
                        return redirect('panel_admin')
                    else:
                        return redirect('panel_funcionario')
                else:
                    messages.error(request, 'Contraseña incorrecta')
                    
            except Usuario.DoesNotExist:
                messages.error(request, 'Usuario no encontrado.')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
                
    else:
        form = LoginForm()
    
    return render(request, 'usuarios/login.html', {'form': form})


def registro_usuario(request):
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            try:
                numero_documento = int(form.cleaned_data['numero_documento'])
                centro = form.cleaned_data['cod_centro_fk']
                coordinacion = form.cleaned_data['cod_coordinacion_fk']
                
                usuario = Usuario.objects.create(
                    id=numero_documento,
                    nombre=form.cleaned_data['nombre'],
                    email=form.cleaned_data['email'],
                    telefono=form.cleaned_data['telefono'],
                    cod_centro_fk=centro,
                    cod_coordinacion_fk=coordinacion,
                    password=make_password(form.cleaned_data['password']),
                    tipo='funcionario'
                )
                
                request.session.flush()
                messages.success(request, '¡Cuenta creada! Inicia sesión.')
                return redirect('login')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
    else:
        form = RegistroUsuarioForm()
    
    return render(request, 'usuarios/registro.html', {'form': form})


@login_required
def perfil_usuario(request):
    try:
        usuario = Usuario.objects.get(id=request.session.get('user_id'))
    except Usuario.DoesNotExist:
        messages.error(request, 'Usuario no encontrado')
        return redirect('login')
    
    if request.method == 'POST':
        form = EditarPerfilForm(request.POST)
        if form.is_valid():
            try:
                password_actual = form.cleaned_data.get('password_actual')
                if password_actual:
                    if not check_password(password_actual, usuario.password):
                        messages.error(request, 'Contraseña incorrecta')
                        return render(request, 'usuarios/perfil.html', {'usuario': usuario, 'form': form})
                
                usuario.email = form.cleaned_data['email']
                usuario.telefono = form.cleaned_data['telefono']
                usuario.cod_centro_fk = form.cleaned_data['cod_centro_fk']
                usuario.cod_coordinacion_fk = form.cleaned_data['cod_coordinacion_fk']
                
                if form.cleaned_data.get('password_nueva'):
                    usuario.password = make_password(form.cleaned_data['password_nueva'])
                
                usuario.save()
                messages.success(request, '¡Perfil actualizado!')
                return redirect('perfil')
                
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
    else:
        form = EditarPerfilForm(initial={
            'email': usuario.email,
            'telefono': usuario.telefono,
            'cod_centro_fk': usuario.cod_centro_fk,
            'cod_coordinacion_fk': usuario.cod_coordinacion_fk
        })
    
    return render(request, 'usuarios/perfil.html', {'usuario': usuario, 'form': form})


def logout_view(request):
    request.session.flush()
    messages.success(request, 'Sesión cerrada')
    return redirect('login')



@login_required
@admin_required
def panel_admin(request):
    try:
        
        total_usuarios = Usuario.objects.count()
        total_funcionarios = Usuario.objects.filter(tipo='funcionario').count()
        total_admins = Usuario.objects.filter(tipo='admin').count()

        
        total_reservas = Reserva.objects.count()
        reservas_pendientes = Reserva.objects.filter(estado='pendiente').count()
        reservas_aprobadas = Reserva.objects.filter(estado='aprobada').count()
        reservas_hoy = Reserva.objects.filter(fecha=datetime.now().date()).count()

       
        total_ambientes = Ambiente.objects.count()
        ambientes_disponibles = Ambiente.objects.filter(estado_amb='disponible').count()

        
        ahora = datetime.now()
        mes_actual = ahora.month
        anio_actual = ahora.year
        reservas_mes = Reserva.objects.filter(
            fecha__month=mes_actual,
            fecha__year=anio_actual
        ).count()

        
        ultimas_reservas = (
            Reserva.objects
            .select_related('cod_usuario_fk', 'cod_ambiente_fk')
            .order_by('-fecha', '-hora_ini')[:6]
        )

        
        fecha_hoy = ahora.date()
        fecha_proxima_semana = fecha_hoy + timedelta(days=7)
        reservas_proximas = (
            Reserva.objects.filter(
                fecha__range=[fecha_hoy, fecha_proxima_semana],
                estado='aprobada'
            )
            .select_related('cod_usuario_fk', 'cod_ambiente_fk')
            .order_by('fecha', 'hora_ini')[:5]
        )

        reservas_por_mes = []
        meses_nombres = []
        for i in range(5, -1, -1):
            fecha = ahora - timedelta(days=30 * i)
            mes = fecha.month
            anio = fecha.year
            count = Reserva.objects.filter(
                fecha__month=mes,
                fecha__year=anio
            ).count()
            reservas_por_mes.append(count)
            meses_nombres.append(fecha.strftime('%b'))

        
        estados_data = {
            'pendientes': reservas_pendientes,
            'aprobadas': reservas_aprobadas,
            'rechazadas': Reserva.objects.filter(estado='rechazada').count(),
        }

        
        ambientes_top = (
            Reserva.objects.values('cod_ambiente_fk__nom_amb')
            .annotate(total=Count('cod_reserva'))
            .order_by('-total')[:5]
        )

       
        usuarios_activos = (
            Reserva.objects.values('cod_usuario_fk__nombre')
            .annotate(total=Count('cod_reserva'))
            .order_by('-total')[:5]
        )

       
        notificaciones = (
            Reserva.objects.filter(estado='pendiente')
            .select_related('cod_usuario_fk', 'cod_ambiente_fk')
            .order_by('-fecha', '-hora_ini')[:10]
        )

        
        alertas = []

        if reservas_pendientes > 0:
            alertas.append({
                'tipo': 'warning',
                'icono': 'fa-clock',
                'titulo': 'Reservas Pendientes',
                'mensaje': f'Tienes {reservas_pendientes} reserva(s) pendiente(s) por aprobar',
                'tiempo': 'Ahora'
            })

        if reservas_hoy > 0:
            alertas.append({
                'tipo': 'info',
                'icono': 'fa-calendar-day',
                'titulo': 'Reservas de Hoy',
                'mensaje': f'Hay {reservas_hoy} reserva(s) programada(s) para hoy',
                'tiempo': 'Hoy'
            })

        ambientes_ocupados = total_ambientes - ambientes_disponibles
        if ambientes_ocupados > 0:
            alertas.append({
                'tipo': 'info',
                'icono': 'fa-door-closed',
                'titulo': 'Ambientes Ocupados',
                'mensaje': f'{ambientes_ocupados} ambiente(s) no disponible(s)',
                'tiempo': '1h'
            })

        context = {
            'total_usuarios': total_usuarios,
            'total_funcionarios': total_funcionarios,
            'total_admins': total_admins,
            'total_reservas': total_reservas,
            'reservas_pendientes': reservas_pendientes,
            'reservas_aprobadas': reservas_aprobadas,
            'reservas_mes': reservas_mes,
            'total_ambientes': total_ambientes,
            'ambientes_disponibles': ambientes_disponibles,

            'ultimas_reservas': ultimas_reservas,
            'reservas_proximas': reservas_proximas,
            'alertas': alertas,

            'reservas_por_mes': reservas_por_mes,
            'meses_nombres': meses_nombres,
            'estados_data': estados_data,
            'ambientes_top': list(ambientes_top),
            'usuarios_activos': list(usuarios_activos),

            
            'notificaciones': notificaciones,
        }

        soportes_abiertos = Soporte.objects.filter(estado='abierto').select_related('usuario').order_by('-creado_en')[:10]
        context['soportes_abiertos'] = soportes_abiertos

        return render(request, 'usuarios/admin_dashboard.html', context)

    except Exception as e:
        messages.error(request, f'Error al cargar dashboard: {str(e)}')
        return render(request, 'usuarios/admin_dashboard.html', {})



@login_required
@funcionario_required
def panel_funcionario(request):
    usuario = Usuario.objects.get(id=request.session['user_id'])
    reservas_asignadas = (
        Reserva.objects
        .filter(cod_usuario_fk=usuario, estado='aprobada')
        .order_by('-fecha', '-hora_ini')[:10]
    )
    reservas_asignadas_count = reservas_asignadas.count()
    context = {
        'reservas_asignadas': reservas_asignadas,
        'reservas_asignadas_count': reservas_asignadas_count,
    }
    return render(request, 'usuarios/funcionario_dashboard.html', context)


@login_required
@admin_required
def centros_list(request):
    from centros.forms import CentroForm
    
    if request.method == 'POST':
        form = CentroForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Centro creado exitosamente')
                return redirect('centros_list')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
    else:
        form = CentroForm()
    
    centros = Centro.objects.all()
    return render(request, 'centros/listar_centros.html', {
        'centros': centros,
        'form': form
    })


@login_required
@admin_required
def coordinaciones_list(request):
    """Gestión de coordinaciones con datos de BD"""
    coordinaciones = Coordinacion.objects.all()
    return render(request, 'coordinaciones/listar_coordinaciones.html', {'coordinaciones': coordinaciones})


@login_required
@admin_required
def reservas_list(request):
    """Gestión de reservas con datos de BD"""
    if Reserva is None:
        messages.error(request, 'Modelo Reserva no encontrado')
        return redirect('panel_admin')
    
    reservas = Reserva.objects.all()
    return render(request, 'reservas/listar_reservas.html', {'reservas': reservas})


@login_required
@admin_required
def usuarios_list(request):
    usuarios = Usuario.objects.all()
    admin_count = Usuario.objects.filter(tipo='admin').count()
    funcionario_count = Usuario.objects.filter(tipo='funcionario').count()
    
    context = {
        'usuarios': usuarios,
        'admin_count': admin_count,
        'funcionario_count': funcionario_count
    }
    return render(request, 'usuarios/usuarios_list.html', context)



@login_required
@admin_required
def calendario_reservas(request):
    try:
        
        mes = int(request.GET.get('mes', datetime.now().month))
        anio = int(request.GET.get('anio', datetime.now().year))
        

        fecha_inicio = datetime(anio, mes, 1)
        if mes == 12:
            fecha_fin = datetime(anio + 1, 1, 1) - __import__('datetime').timedelta(days=1)
        else:
            fecha_fin = datetime(anio, mes + 1, 1) - __import__('datetime').timedelta(days=1)
        
        reservas = Reserva.objects.filter(
            fecha__range=[fecha_inicio.date(), fecha_fin.date()]
        ).values('fecha', 'cod_ambiente_fk__nom_amb', 'cod_usuario_fk__nombre', 'estado').order_by('fecha')
        
        
        import json
        reservas_por_fecha = {}
        for res in reservas:
            fecha = str(res['fecha'])
            if fecha not in reservas_por_fecha:
                reservas_por_fecha[fecha] = []
            
            reservas_por_fecha[fecha].append({
                'ambiente': res['cod_ambiente_fk__nom_amb'],
                'usuario': res['cod_usuario_fk__nombre'],
                'estado': res['estado']
            })
        
        
        total_reservas = len(reservas)
        aprobadas = reservas.filter(estado='aprobada').count()
        pendientes = reservas.filter(estado='pendiente').count()
        rechazadas = reservas.filter(estado='rechazada').count()
        
        context = {
            'mes': mes,
            'anio': anio,
            'mes_nombre': ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                          'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][mes - 1],
            'reservas_por_fecha': json.dumps(reservas_por_fecha),
            'total_reservas': total_reservas,
            'aprobadas': aprobadas,
            'pendientes': pendientes,
            'rechazadas': rechazadas,
        }
        
        return render(request, 'usuarios/calendario_reservas.html', context)
    
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('panel_admin')


TIPO_PROBLEMA_CHOICES = [
    ('sonido', 'Sonido'),
    ('inventario', 'Inventario (equipos, muebles, etc.)'),
    ('iluminacion', 'Iluminación'),
    ('aseo', 'Aseo / limpieza'),
    ('infraestructura', 'Infraestructura'),
    ('conectividad', 'Conectividad (red / energía)'),
    ('otros', 'Otros'),
]


@login_required
@admin_required
def reportes_view(request):
    reporte_data = None
    tipo_reporte = request.GET.get('tipo', 'reservas')
    formato = request.GET.get('formato', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    estado_filtro = request.GET.get('estado', '').strip()
    ambiente_filtro = request.GET.get('ambiente', '').strip()
    usuario_filtro = request.GET.get('usuario', '').strip()
    tipo_problema_filtro = request.GET.get('tipo_problema', '').strip()

    ambientes_list = Ambiente.objects.all()
    usuarios_list = Usuario.objects.filter(tipo='funcionario')

    try:
        filtros = Q()

        if fecha_inicio:
            try:
                fecha_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                filtros &= Q(fecha__gte=fecha_ini)
            except Exception:
                pass

        if fecha_fin:
            try:
                fecha_f = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                filtros &= Q(fecha__lte=fecha_f)
            except Exception:
                pass

        if estado_filtro:
            filtros &= Q(estado=estado_filtro)

        if ambiente_filtro:
            filtros &= Q(cod_ambiente_fk=ambiente_filtro)

        if usuario_filtro:
            filtros &= Q(cod_usuario_fk=usuario_filtro)

        if tipo_problema_filtro:
            filtros &= Q(tipos_problema__icontains=tipo_problema_filtro)

        if tipo_reporte == 'reservas':
            qs_res = Reserva.objects.filter(filtros)
            reporte_data = list(
                qs_res.values(
                    'cod_reserva',
                    'titulo_reserva',
                    'tipo_reserva',
                    'fecha',
                    'hora_ini',
                    'hora_fin',
                    'motivo',
                    'num_personas',
                    'estado',
                    'tipos_problema',
                    'cod_usuario_fk__nombre',
                    'cod_usuario_fk__email',
                    'cod_ambiente_fk__nom_amb',
                    'cod_ambiente_fk__capacidad_amb',
                ).order_by('-fecha')
            )

        elif tipo_reporte == 'ambientes':
            qs_amb = Ambiente.objects.all()
            reporte_data = list(
                qs_amb.annotate(
                    total_reservas=Count('reserva', filter=filtros)
                ).values(
                    'cod_amb',
                    'nom_amb',
                    'capacidad_amb',
                    'piso_amb',
                    'info_amb',
                    'estado_amb',
                    'total_reservas',
                ).order_by('-total_reservas')
            )

        elif tipo_reporte == 'funcionarios':
            qs_usr = Usuario.objects.filter(tipo='funcionario')
            reporte_data = list(
                qs_usr.annotate(
                    total_reservas=Count('cod_usuario_fk', filter=filtros)
                ).values(
                    'id',
                    'nombre',
                    'email',
                    'telefono',
                    'cod_centro_fk__nom_centro',
                    'total_reservas',
                ).order_by('-total_reservas')
            )

        if formato == 'excel' and reporte_data:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"

            header_fill = PatternFill(start_color="1d5a3d", end_color="1d5a3d", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'),
            )

            if tipo_reporte == 'reservas':
                headers = [
                    'Código', 'Título', 'Tipo', 'Fecha', 'Hora Inicio', 'Hora Fin',
                    'Motivo', 'Personas', 'Estado', 'Tipo problema',
                    'Funcionario', 'Email', 'Ambiente', 'Capacidad',
                ]
                col_widths = [10, 18, 12, 12, 12, 12, 25, 8, 12, 18, 18, 22, 18, 10]
            elif tipo_reporte == 'ambientes':
                headers = ['Código', 'Nombre', 'Capacidad', 'Piso', 'Información', 'Estado', 'Total Reservas']
                col_widths = [10, 20, 12, 8, 25, 15, 15]
            else:
                headers = ['Documento', 'Nombre', 'Email', 'Teléfono', 'Centro', 'Total Reservas']
                col_widths = [15, 20, 25, 15, 20, 15]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

            for idx, row in enumerate(reporte_data, 2):
                if tipo_reporte == 'reservas':
                    datos = [
                        row.get('cod_reserva'),
                        row.get('titulo_reserva'),
                        row.get('tipo_reserva'),
                        row.get('fecha'),
                        row.get('hora_ini'),
                        row.get('hora_fin'),
                        row.get('motivo'),
                        row.get('num_personas'),
                        row.get('estado'),
                        row.get('tipos_problema'),
                        row.get('cod_usuario_fk__nombre'),
                        row.get('cod_usuario_fk__email'),
                        row.get('cod_ambiente_fk__nom_amb'),
                        row.get('cod_ambiente_fk__capacidad_amb'),
                    ]
                elif tipo_reporte == 'ambientes':
                    datos = [
                        row.get('cod_amb'),
                        row.get('nom_amb'),
                        row.get('capacidad_amb'),
                        row.get('piso_amb'),
                        row.get('info_amb'),
                        row.get('estado_amb'),
                        row.get('total_reservas'),
                    ]
                else:
                    datos = [
                        row.get('id'),
                        row.get('nombre'),
                        row.get('email'),
                        row.get('telefono'),
                        row.get('cod_centro_fk__nom_centro'),
                        row.get('total_reservas'),
                    ]

                for col, dato in enumerate(datos, 1):
                    cell = ws.cell(row=idx, column=col)
                    cell.value = dato
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            for idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            fecha_texto = f"_{fecha_inicio}_a_{fecha_fin}" if fecha_inicio and fecha_fin else ""
            response['Content-Disposition'] = (
                f'attachment; filename="Reporte_{tipo_reporte}{fecha_texto}_{datetime.now().strftime("%d-%m-%Y_%H%M%S")}.xlsx"'
            )
            wb.save(response)
            return response

        elif formato == 'pdf' and reporte_data:
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors

            buffer = BytesIO()
            pagesize = landscape(A4)
            doc = SimpleDocTemplate(
                buffer,
                pagesize=pagesize,
                rightMargin=20,
                leftMargin=20,
                topMargin=20,
                bottomMargin=20,
            )

            elements = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontSize=16,
                textColor=colors.HexColor('#1d5a3d'),
                spaceAfter=12,
                alignment=1,
                fontName='Helvetica-Bold',
            )

            title = Paragraph(f"Reporte de {tipo_reporte.capitalize()}", title_style)
            elements.append(title)

            info_style = ParagraphStyle(
                'Info',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.grey,
                spaceAfter=8,
            )

            info_text = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            if fecha_inicio or fecha_fin:
                info_text += f" | Período: {fecha_inicio or 'Inicio'} a {fecha_fin or 'Fin'}"

            info = Paragraph(info_text, info_style)
            elements.append(info)
            elements.append(Spacer(1, 0.15 * inch))

            if tipo_reporte == 'reservas':
                data = [
                    ['Código', 'Título', 'Tipo', 'Fecha', 'Hora Ini', 'Hora Fin',
                     'Pers.', 'Estado', 'Tipo prob.', 'Func.', 'Amb.']
                ]
                for row in reporte_data:
                    data.append([
                        str(row.get('cod_reserva', '')),
                        str(row.get('titulo_reserva', ''))[:12],
                        str(row.get('tipo_reserva', '')),
                        str(row.get('fecha', '')),
                        str(row.get('hora_ini', '')),
                        str(row.get('hora_fin', '')),
                        str(row.get('num_personas', '')),
                        str(row.get('estado', '')),
                        str(row.get('tipos_problema', ''))[:12],
                        str(row.get('cod_usuario_fk__nombre', ''))[:12],
                        str(row.get('cod_ambiente_fk__nom_amb', ''))[:12],
                    ])
                col_widths = [
                    0.6 * inch, 0.9 * inch, 0.8 * inch, 0.8 * inch, 0.7 * inch,
                    0.7 * inch, 0.6 * inch, 0.8 * inch, 0.9 * inch, 1 * inch, 1 * inch,
                ]

            elif tipo_reporte == 'ambientes':
                data = [['Código', 'Nombre', 'Capacidad', 'Piso', 'Información', 'Estado', 'Total Res']]
                for row in reporte_data:
                    data.append([
                        str(row.get('cod_amb', '')),
                        str(row.get('nom_amb', '')),
                        str(row.get('capacidad_amb', '')),
                        str(row.get('piso_amb', '')),
                        str(row.get('info_amb', ''))[:12],
                        str(row.get('estado_amb', '')),
                        str(row.get('total_reservas', '')),
                    ])
                col_widths = [0.7 * inch, 1.2 * inch, 0.8 * inch, 0.6 * inch, 1 * inch, 1 * inch, 0.8 * inch]

            else:
                data = [['Documento', 'Nombre', 'Email', 'Teléfono', 'Centro', 'Total Res']]
                for row in reporte_data:
                    data.append([
                        str(row.get('id', '')),
                        str(row.get('nombre', '')),
                        str(row.get('email', '')),
                        str(row.get('telefono', '')),
                        str(row.get('cod_centro_fk__nom_centro', ''))[:12],
                        str(row.get('total_reservas', '')),
                    ])
                col_widths = [1 * inch, 1.2 * inch, 1.5 * inch, 1 * inch, 1.2 * inch, 0.8 * inch]

            tabla = Table(data, colWidths=col_widths)
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1d5a3d')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))

            elements.append(tabla)
            doc.build(elements)

            pdf_content = buffer.getvalue()
            buffer.close()

            response = HttpResponse(pdf_content, content_type='application/pdf')
            fecha_nombre = f"_{fecha_inicio}_a_{fecha_fin}" if fecha_inicio and fecha_fin else ""
            filename = f"Reporte_{tipo_reporte}{fecha_nombre}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        elif not reporte_data and formato:
            messages.warning(request, 'No hay datos para descargar con los filtros aplicados')

    except Exception as e:
        messages.error(request, f'Error: {str(e)}')

    context = {
        'reporte_data': reporte_data,
        'tipo_reporte': tipo_reporte,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'estado_filtro': estado_filtro,
        'ambiente_filtro': ambiente_filtro,
        'usuario_filtro': usuario_filtro,
        'tipo_problema_filtro': tipo_problema_filtro,
        'ambientes_list': ambientes_list,
        'usuarios_list': usuarios_list,
        'estados_choices': Reserva.ESTADO_CHOICES,
        'TIPO_PROBLEMA_CHOICES': TIPO_PROBLEMA_CHOICES,
    }
    return render(request, 'usuarios/reportes.html', context)

@login_required
@admin_required
def carga_masiva_usuarios(request):
    """Vista para cargar usuarios de forma masiva"""
    from .forms import CargaMasivausuariosForm
    
    resultado = {
        'exitosos': 0,
        'errores': 0,
        'detalles': []
    }
    
    if request.method == 'POST':
        form = CargaMasivausuariosForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            
            try:
                # Detectar tipo de archivo
                nombre_archivo = archivo.name.lower()
                
                if nombre_archivo.endswith('.csv'):
                    df = pd.read_csv(archivo)
                elif nombre_archivo.endswith(('.xlsx', '.xls')):
                    df = pd.read_excel(archivo)
                else:
                    messages.error(request, 'Formato de archivo no válido. Use CSV o Excel.')
                    return render(request, 'usuarios/carga_masiva.html', {'form': form})
                
                # Procesar cada fila
                for idx, row in df.iterrows():
                    fila_numero = idx + 2  # +2 porque Excel empieza en 1 y hay encabezado
                    
                    try:
                        # Extraer datos de forma flexible
                        numero_documento = str(row.get('numero_documento') or row.get('id') or row.get('documento')).strip()
                        nombre = str(row.get('nombre') or row.get('name')).strip()
                        email = str(row.get('email') or row.get('correo')).strip()
                        telefono = str(row.get('telefono') or row.get('phone') or '').strip()
                        tipo = str(row.get('tipo') or row.get('type') or 'funcionario').strip().lower()
                        
                        # Centro - Por código o por nombre
                        centro_valor = str(row.get('centro') or row.get('center') or '').strip()
                        centro = None
                        
                        if centro_valor:
                            # Intentar por código
                            try:
                                centro = Centro.objects.get(cod_centro=int(centro_valor))
                            except (ValueError, Centro.DoesNotExist):
                                # Intentar por nombre
                                centro = Centro.objects.filter(nom_centro__icontains=centro_valor).first()
                                if not centro:
                                    raise ValueError(f"Centro no encontrado: {centro_valor}")
                        
                        # Coordinación - Por código o por nombre
                        coordinacion_valor = str(row.get('coordinacion') or row.get('coordination') or '').strip()
                        coordinacion = None
                        
                        if coordinacion_valor:
                            # Intentar por código
                            try:
                                coordinacion = Coordinacion.objects.get(cod_coordinacion=int(coordinacion_valor))
                            except (ValueError, Coordinacion.DoesNotExist):
                                # Intentar por nombre
                                coordinacion = Coordinacion.objects.filter(nom_coordinacion__icontains=coordinacion_valor).first()
                                if not coordinacion:
                                    raise ValueError(f"Coordinación no encontrada: {coordinacion_valor}")
                        
                        # Contraseña
                        password = str(row.get('password') or row.get('contraseña') or 'sena123').strip()
                        
                        # Validaciones
                        if not numero_documento:
                            raise ValueError("Documento requerido")
                        if not nombre:
                            raise ValueError("Nombre requerido")
                        if not email:
                            raise ValueError("Email requerido")
                        
                        # Verificar si ya existe
                        if Usuario.objects.filter(id=numero_documento).exists():
                            raise ValueError(f"Usuario con documento {numero_documento} ya existe")
                        
                        if Usuario.objects.filter(email=email).exists():
                            raise ValueError(f"Email {email} ya está registrado")
                        
                        # Validar tipo
                        if tipo not in ['admin', 'funcionario']:
                            tipo = 'funcionario'
                        
                        # Crear usuario
                        Usuario.objects.create(
                            id=numero_documento,
                            nombre=nombre,
                            email=email,
                            telefono=telefono,
                            password=make_password(password),
                            tipo=tipo,
                            cod_centro_fk=centro,
                            cod_coordinacion_fk=coordinacion
                        )
                        
                        resultado['exitosos'] += 1
                        resultado['detalles'].append({
                            'fila': fila_numero,
                            'estado': 'éxito',
                            'documento': numero_documento,
                            'nombre': nombre,
                            'mensaje': 'Usuario creado exitosamente'
                        })
                        
                    except Exception as e:
                        resultado['errores'] += 1
                        resultado['detalles'].append({
                            'fila': fila_numero,
                            'estado': 'error',
                            'documento': numero_documento if 'numero_documento' in locals() else 'N/A',
                            'nombre': nombre if 'nombre' in locals() else 'N/A',
                            'mensaje': str(e)
                        })
                
    
                if resultado['exitosos'] > 0:
                    messages.success(request, f"{resultado['exitosos']} usuarios creados exitosamente")
                if resultado['errores'] > 0:
                    messages.warning(request, f"{resultado['errores']} errores durante la carga")
                
                context = {
                    'form': form,
                    'resultado': resultado,
                    'total': resultado['exitosos'] + resultado['errores']
                }
                return render(request, 'usuarios/carga_masiva.html', context)
                
            except Exception as e:
                messages.error(request, f'Error al procesar archivo: {str(e)}')
                return render(request, 'usuarios/carga_masiva.html', {'form': form})
    else:
        form = CargaMasivausuariosForm()
    
    return render(request, 'usuarios/carga_masiva.html', {'form': form})
